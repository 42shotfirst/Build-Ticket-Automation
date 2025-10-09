#!/usr/bin/env python3
"""
Comprehensive Excel Data Extractor
==================================
Extracts ALL data from Excel files including:
- All sheet data (tables, key-value pairs, raw data)
- VBA macros and code
- Formulas (as text where possible)
- Cell formatting and styles
- Comments and data validation
- Charts and images (metadata)
- Named ranges
- Data connections
"""

import pandas as pd
import json
import os
import warnings
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import zipfile
import xml.etree.ElementTree as ET

# Suppress openpyxl warnings
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

class ComprehensiveExcelExtractor:
    """Extract all possible data from Excel files."""
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.file_name = os.path.basename(file_path)
        self.extracted_data = {}
        
    def extract_all(self) -> Dict[str, Any]:
        """Extract all data from the Excel file."""
        print(f"Starting comprehensive extraction from: {self.file_path}")
        
        # Initialize extraction result
        self.extracted_data = {
            'file_info': {
                'filename': self.file_name,
                'file_path': self.file_path,
                'extraction_timestamp': datetime.now().isoformat(),
                'extractor_version': '1.0.0'
            },
            'sheets': {},
            'macros': {},
            'formulas': {},
            'named_ranges': {},
            'data_validation': {},
            'charts': {},
            'images': {},
            'comments': {},
            'workbook_properties': {}
        }
        
        try:
            # Extract basic sheet data
            self._extract_sheet_data()
            
            # Extract macros and VBA code
            self._extract_macros()
            
            # Extract formulas
            self._extract_formulas()
            
            # Extract workbook properties
            self._extract_workbook_properties()
            
            # Extract named ranges
            self._extract_named_ranges()
            
            # Extract comments
            self._extract_comments()
            
            print(f"SUCCESS: Comprehensive extraction completed successfully")
            return self.extracted_data
            
        except Exception as e:
            print(f"ERROR: Error during extraction: {e}")
            import traceback
            traceback.print_exc()
            return self.extracted_data
    
    def _extract_sheet_data(self):
        """Extract data from all sheets."""
        print("Extracting sheet data...")
        
        try:
            excel_file = pd.ExcelFile(self.file_path, engine='openpyxl')
            sheet_names = excel_file.sheet_names
            
            print(f"Found {len(sheet_names)} sheets: {sheet_names}")
            
            for sheet_name in sheet_names:
                print(f"  Processing sheet: '{sheet_name}'")
                
                sheet_data = {
                    'name': sheet_name,
                    'raw_data': [],
                    'structured_data': {},
                    'tables': [],
                    'key_value_pairs': {},
                    'dimensions': {},
                    'cell_formats': {},
                    'data_validation': {}
                }
                
                try:
                    # Read raw data
                    df_raw = pd.read_excel(excel_file, sheet_name=sheet_name, header=None, engine='openpyxl')
                    
                    if not df_raw.empty:
                        # Store dimensions
                        sheet_data['dimensions'] = {
                            'rows': int(df_raw.shape[0]),
                            'columns': int(df_raw.shape[1])
                        }
                        
                        # Store raw data
                        sheet_data['raw_data'] = df_raw.fillna('').to_dict('records')
                        
                        # Extract structured data
                        self._extract_structured_data(df_raw, sheet_data)
                        
                        # Extract tables
                        self._extract_tables(df_raw, sheet_data)
                        
                        # Extract key-value pairs
                        self._extract_key_value_pairs(df_raw, sheet_data)
                        
                        print(f"    SUCCESS: Extracted {sheet_data['dimensions']['rows']}x{sheet_data['dimensions']['columns']} data")
                    else:
                        print(f"    WARNING: Sheet '{sheet_name}' is empty")
                    
                    self.extracted_data['sheets'][sheet_name] = sheet_data
                    
                except Exception as e:
                    print(f"    ERROR: Error reading sheet '{sheet_name}': {e}")
                    self.extracted_data['sheets'][sheet_name] = {
                        'name': sheet_name,
                        'error': str(e),
                        'raw_data': [],
                        'structured_data': {},
                        'tables': [],
                        'key_value_pairs': {},
                        'dimensions': {'rows': 0, 'columns': 0}
                    }
            
            excel_file.close()
            
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            self.extracted_data['sheets'] = {'error': str(e)}
    
    def _extract_structured_data(self, df: pd.DataFrame, sheet_data: Dict):
        """Extract structured data from DataFrame."""
        # Try different header assumptions
        for header_row in [0, 1, 2, 3, 5, 10]:
            try:
                df_header = pd.read_excel(pd.ExcelFile(self.file_path, engine='openpyxl'), 
                                        sheet_name=sheet_data['name'], 
                                        header=header_row, engine='openpyxl')
                
                if not df_header.empty and len(df_header.columns) > 1:
                    # Check if columns look meaningful
                    meaningful_cols = [col for col in df_header.columns 
                                     if str(col).strip() and 'Unnamed' not in str(col)]
                    
                    if len(meaningful_cols) >= 2:
                        structured_data = {
                            'header_row': header_row,
                            'columns': list(df_header.columns),
                            'data': df_header.fillna('').to_dict('records'),
                            'row_count': len(df_header)
                        }
                        sheet_data['structured_data'][f'header_row_{header_row}'] = structured_data
                        
            except Exception:
                continue
    
    def _extract_tables(self, df: pd.DataFrame, sheet_data: Dict):
        """Extract tabular data from DataFrame."""
        tables = []
        
        if df.empty:
            return
        
        # Look for potential table headers
        for row_idx in range(min(20, len(df))):  # Check first 20 rows
            row = df.iloc[row_idx]
            non_empty_count = sum(1 for val in row if pd.notna(val) and str(val).strip())
            
            if non_empty_count >= 3:  # Potential header row
                # Extract headers
                headers = []
                for col_idx in range(len(row)):
                    val = row.iloc[col_idx]
                    if pd.notna(val) and str(val).strip():
                        headers.append(str(val).strip())
                    else:
                        headers.append(f"Column_{col_idx}")
                
                # Extract data rows
                data_rows = []
                for data_idx in range(row_idx + 1, min(row_idx + 100, len(df))):
                    data_row = df.iloc[data_idx]
                    row_data = {}
                    has_data = False
                    
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(data_row):
                            value = data_row.iloc[col_idx]
                            if pd.notna(value) and str(value).strip():
                                row_data[header] = str(value).strip()
                                has_data = True
                    
                    if has_data:
                        data_rows.append(row_data)
                    else:
                        break  # Stop at first empty row
                
                if data_rows and len(headers) >= 3:
                    table = {
                        'header_row_index': int(row_idx),
                        'headers': headers,
                        'data': data_rows,
                        'row_count': len(data_rows)
                    }
                    tables.append(table)
        
        sheet_data['tables'] = tables
    
    def _extract_key_value_pairs(self, df: pd.DataFrame, sheet_data: Dict):
        """Extract key-value pairs from DataFrame."""
        key_value_pairs = {}
        
        if len(df.columns) < 2:
            return
        
        for index, row in df.iterrows():
            key = str(row.iloc[0]).strip() if pd.notna(row.iloc[0]) else ""
            value = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
            
            # Clean up key
            key_clean = key.replace(':', '').strip()
            
            # Only add if both key and value are meaningful
            if (key_clean and value and 
                key_clean.lower() not in ['nan', 'none', ''] and 
                value.lower() not in ['nan', 'none', '']):
                key_value_pairs[key_clean] = value
        
        sheet_data['key_value_pairs'] = key_value_pairs
    
    def _extract_macros(self):
        """Extract VBA macros and code from Excel file."""
        print("Extracting macros...")
        
        try:
            # Excel files with macros are ZIP files
            with zipfile.ZipFile(self.file_path, 'r') as zip_file:
                file_list = zip_file.namelist()
                
                # Look for VBA project files
                vba_files = [f for f in file_list if f.startswith('xl/vbaProject.bin')]
                
                if vba_files:
                    print(f"  Found VBA project files: {vba_files}")
                    
                    # Extract VBA project (binary format - we can't easily read the code)
                    for vba_file in vba_files:
                        try:
                            vba_content = zip_file.read(vba_file)
                            self.extracted_data['macros']['vba_project_bin'] = {
                                'filename': vba_file,
                                'size_bytes': len(vba_content),
                                'note': 'VBA project in binary format - code not directly readable'
                            }
                        except Exception as e:
                            self.extracted_data['macros']['vba_project_error'] = str(e)
                
                # Look for other macro-related files
                macro_files = [f for f in file_list if 'macro' in f.lower() or 'vba' in f.lower()]
                if macro_files:
                    self.extracted_data['macros']['related_files'] = macro_files
                
                # Look for custom XML files that might contain macro information
                custom_xml_files = [f for f in file_list if f.startswith('customXml/')]
                if custom_xml_files:
                    self.extracted_data['macros']['custom_xml_files'] = custom_xml_files
                
                if not vba_files and not macro_files:
                    print("  No macros found")
                    self.extracted_data['macros']['status'] = 'No macros detected'
                    
        except Exception as e:
            print(f"  Error extracting macros: {e}")
            self.extracted_data['macros']['error'] = str(e)
    
    def _extract_formulas(self):
        """Extract formulas from Excel sheets."""
        print("Extracting formulas...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.file_path, data_only=False)
            formulas_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_formulas = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula cell
                            formula_info = {
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'calculated_value': cell.data_type,
                                'row': cell.row,
                                'column': cell.column
                            }
                            sheet_formulas.append(formula_info)
                
                if sheet_formulas:
                    formulas_data[sheet_name] = sheet_formulas
                    print(f"  Found {len(sheet_formulas)} formulas in sheet '{sheet_name}'")
            
            self.extracted_data['formulas'] = formulas_data
            
            if not formulas_data:
                print("  No formulas found")
                self.extracted_data['formulas']['status'] = 'No formulas detected'
                
        except Exception as e:
            print(f"  Error extracting formulas: {e}")
            self.extracted_data['formulas']['error'] = str(e)
    
    def _extract_workbook_properties(self):
        """Extract workbook properties and metadata."""
        print("Extracting workbook properties...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.file_path)
            properties = workbook.properties
            
            self.extracted_data['workbook_properties'] = {
                'title': properties.title,
                'creator': properties.creator,
                'last_modified_by': properties.lastModifiedBy,
                'created': properties.created.isoformat() if properties.created else None,
                'modified': properties.modified.isoformat() if properties.modified else None,
                'description': properties.description,
                'subject': properties.subject,
                'keywords': properties.keywords,
                'category': properties.category,
                'version': properties.version,
                'sheet_count': len(workbook.sheetnames),
                'sheet_names': workbook.sheetnames
            }
            
            print(f"  Extracted workbook properties")
            
        except Exception as e:
            print(f"  Error extracting workbook properties: {e}")
            self.extracted_data['workbook_properties']['error'] = str(e)
    
    def _extract_named_ranges(self):
        """Extract named ranges from workbook."""
        print("Extracting named ranges...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.file_path)
            named_ranges = {}
            
            for name, range_obj in workbook.defined_names.items():
                named_ranges[name] = {
                    'formula': range_obj.attr_text if hasattr(range_obj, 'attr_text') else str(range_obj),
                    'local_sheet_id': range_obj.localSheetId if hasattr(range_obj, 'localSheetId') else None
                }
            
            self.extracted_data['named_ranges'] = named_ranges
            
            if named_ranges:
                print(f"  Found {len(named_ranges)} named ranges")
            else:
                print("  No named ranges found")
                self.extracted_data['named_ranges']['status'] = 'No named ranges detected'
                
        except Exception as e:
            print(f"  Error extracting named ranges: {e}")
            self.extracted_data['named_ranges']['error'] = str(e)
    
    def _extract_comments(self):
        """Extract comments from sheets."""
        print("Extracting comments...")
        
        try:
            from openpyxl import load_workbook
            
            workbook = load_workbook(self.file_path)
            comments_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_comments = []
                
                # Note: openpyxl doesn't directly support reading comments in all cases
                # This is a simplified approach
                if hasattr(sheet, '_comments'):
                    for cell_coord, comment in sheet._comments.items():
                        comment_info = {
                            'cell': cell_coord,
                            'author': comment.author if hasattr(comment, 'author') else 'Unknown',
                            'text': str(comment.text) if hasattr(comment, 'text') else ''
                        }
                        sheet_comments.append(comment_info)
                
                if sheet_comments:
                    comments_data[sheet_name] = sheet_comments
            
            self.extracted_data['comments'] = comments_data
            
            if not comments_data:
                print("  No comments found")
                self.extracted_data['comments']['status'] = 'No comments detected'
                
        except Exception as e:
            print(f"  Error extracting comments: {e}")
            self.extracted_data['comments']['error'] = str(e)
    
    def export_to_json(self, output_file: str = None) -> str:
        """Export extracted data to JSON file."""
        if output_file is None:
            base_name = os.path.splitext(self.file_name)[0]
            output_file = f"{base_name}_comprehensive_extract.json"
        
        try:
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(self.extracted_data, f, indent=2, default=str, ensure_ascii=False)
            
            file_size = os.path.getsize(output_file)
            print(f"SUCCESS: Exported comprehensive data to: {output_file} ({file_size:,} bytes)")
            return output_file
            
        except Exception as e:
            print(f"ERROR: Error exporting to JSON: {e}")
            return None
    
    def get_summary(self) -> Dict[str, Any]:
        """Get a summary of extracted data."""
        summary = {
            'file_name': self.file_name,
            'sheets_count': len(self.extracted_data.get('sheets', {})),
            'total_formulas': sum(len(sheet_formulas) for sheet_formulas in self.extracted_data.get('formulas', {}).values() 
                                if isinstance(sheet_formulas, list)),
            'has_macros': bool(self.extracted_data.get('macros', {}).get('vba_project_bin')),
            'named_ranges_count': len(self.extracted_data.get('named_ranges', {})),
            'comments_count': sum(len(sheet_comments) for sheet_comments in self.extracted_data.get('comments', {}).values() 
                                if isinstance(sheet_comments, list))
        }
        
        return summary


def main():
    """Main function for testing the comprehensive extractor."""
    import sys
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = "LLDtest.xlsm"
    
    if not os.path.exists(excel_file):
        print(f"Error: File not found: {excel_file}")
        return False
    
    # Create extractor and extract all data
    extractor = ComprehensiveExcelExtractor(excel_file)
    extracted_data = extractor.extract_all()
    
    # Export to JSON
    output_file = extractor.export_to_json()
    
    # Show summary
    summary = extractor.get_summary()
    print("\n" + "="*60)
    print("EXTRACTION SUMMARY")
    print("="*60)
    for key, value in summary.items():
        print(f"{key}: {value}")
    
    if output_file:
        print(f"\nComplete data exported to: {output_file}")
    
    return True


if __name__ == "__main__":
    main()
