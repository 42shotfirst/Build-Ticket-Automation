"""
ACR NRS Sheet Extractor

Extracts data from the ACR NRS (Azure Container Registry Network Rule Set) sheet.
This sheet has a flat table structure with headers in the first row.

Structure:
- Row 1: Headers (action, ip_range, etc.)
- Row 2+: ACR network rule data
"""

import openpyxl
from typing import Dict, Any, List, Optional


class ACRExtractor:
    """Extracts Azure Container Registry network rules from ACR NRS sheet."""

    def __init__(self, excel_path: str):
        """
        Initialize the ACR extractor.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.wb = None
        self.ws = None

    def extract(self) -> Dict[str, Any]:
        """
        Extract all ACR network rules as flat data.

        Returns:
            Dictionary with ACR network rules and metadata
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['ACR NRS']

        data = {
            'network_rules': self._extract_network_rules(),
            'headers': self._extract_headers(),
        }

        self.wb.close()
        return data

    def _extract_headers(self) -> List[str]:
        """
        Extract column headers from first row.

        Returns:
            List of header names
        """
        headers = []

        if self.ws.max_row < 1:
            return headers

        # Read first row for headers
        for col_idx in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(1, col_idx).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                # If no header, stop reading columns
                break

        return headers

    def _extract_network_rules(self) -> List[Dict[str, Any]]:
        """
        Extract all ACR network rules as flat data.

        Reads the entire table structure and converts each row to a dictionary
        using the headers as keys.

        Returns:
            List of network rule dictionaries
        """
        rules = []

        if self.ws.max_row < 2:
            return rules

        # Get headers from first row
        headers = self._extract_headers()

        if not headers:
            return rules

        # Read data rows (starting from row 2)
        for row_idx in range(2, self.ws.max_row + 1):
            rule = {}
            has_data = False

            for col_idx, header in enumerate(headers, start=1):
                if col_idx > self.ws.max_column:
                    break

                cell_value = self.ws.cell(row_idx, col_idx).value

                if cell_value is not None and str(cell_value).strip():
                    rule[header] = self._parse_cell_value(str(cell_value).strip())
                    has_data = True
                else:
                    rule[header] = None

            # Only add rule if it has at least some data
            if has_data:
                # Clean up the rule - remove None values for cleaner output
                cleaned_rule = {k: v for k, v in rule.items() if v is not None}
                if cleaned_rule:
                    rules.append(cleaned_rule)

        return rules

    def _parse_cell_value(self, value: str) -> Any:
        """
        Parse cell value and convert to appropriate type.

        Args:
            value: String value from cell

        Returns:
            Parsed value (bool, list, or string)
        """
        # Check for boolean values
        if value.lower() in ['true', 'yes', 'allow']:
            return True
        if value.lower() in ['false', 'no', 'deny']:
            return False

        # Check for list values (comma-separated)
        if ',' in value:
            return [item.strip() for item in value.split(',')]

        # Return as string
        return value

    def get_network_rules_for_terraform(self) -> List[Dict[str, Any]]:
        """
        Get ACR network rules formatted for terraform.

        Returns:
            List of rules with terraform-compatible field names
        """
        data = self.extract()
        rules = data.get('network_rules', [])

        terraform_rules = []

        for rule in rules:
            # Map Excel column names to terraform field names
            tf_rule = {
                'action': rule.get('action', 'Allow'),
                'ip_address_or_range': rule.get('ip_range', rule.get('ip_address_or_range', '')),
            }

            # Only add if we have an IP range
            if tf_rule['ip_address_or_range']:
                terraform_rules.append(tf_rule)

        return terraform_rules
