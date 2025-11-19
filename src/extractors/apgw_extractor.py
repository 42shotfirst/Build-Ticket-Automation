"""
APGW Sheet Extractor

Extracts data from the APGW (Application Gateway) sheet which has the structure:
- Column A: Label/Terraform Variable (variable name)
- Column B: Value (actual value to use)

Note: Different from Build_ENV which has TF Variable in Column B!
This sheet has TF Variable in Column A and Value in Column B.
"""

import openpyxl
from typing import Dict, Any, List, Optional


class APGWExtractor:
    """Extracts Application Gateway configuration from APGW sheet."""

    def __init__(self, excel_path: str):
        """
        Initialize the APGW extractor.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.wb = None
        self.ws = None

    def extract(self) -> Dict[str, Any]:
        """
        Extract all Application Gateway configuration.

        Returns:
            Dictionary with APGW configuration organized by component type
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['APGW']

        data = {
            'backend_address_pools': self._extract_backend_pools(),
            'backend_http_settings': self._extract_backend_http_settings(),
            'frontend_ip_configurations': self._extract_frontend_ip_configs(),
            'frontend_ports': self._extract_frontend_ports(),
            'http_listeners': self._extract_http_listeners(),
            'request_routing_rules': self._extract_routing_rules(),
            'probes': self._extract_probes(),
            'all_variables': self._extract_all_variables(),
        }

        self.wb.close()
        return data

    def _extract_all_variables(self) -> Dict[str, str]:
        """
        Extract all terraform variables and values.

        Reads Column A (terraform variable) and Column B (value).

        Returns:
            Dictionary mapping variable names to values
        """
        variables = {}

        for row_idx in range(1, self.ws.max_row + 1):
            col_a = self.ws.cell(row_idx, 1).value  # Terraform Variable
            col_b = self.ws.cell(row_idx, 2).value  # Value

            if col_a and col_b:
                tf_var = str(col_a).strip()
                value = str(col_b).strip()

                # Skip header rows
                if value and value not in ['Value', '']:
                    variables[tf_var] = value

        return variables

    def _find_section_data(self, section_name: str) -> List[Dict[str, Any]]:
        """
        Find all instances of a section and extract their data.

        A section is identified by having the section name in Column A
        with 'Value' in Column B.

        Args:
            section_name: Name of section to find (e.g., 'backend_address_pool')

        Returns:
            List of dictionaries, one for each section instance
        """
        sections = []
        current_section = None

        for row_idx in range(1, self.ws.max_row + 1):
            col_a = self.ws.cell(row_idx, 1).value
            col_b = self.ws.cell(row_idx, 2).value

            # Check if this is a section header
            if col_a and str(col_a).strip() == section_name and col_b == 'Value':
                # Save previous section if exists
                if current_section:
                    sections.append(current_section)

                # Start new section
                current_section = {}
                continue

            # If we're in a section, collect data
            if current_section is not None:
                if col_a:
                    var_name = str(col_a).strip()

                    # Stop if we hit another major section or empty row pattern
                    if col_b == 'Value' and var_name != section_name:
                        # This is a new section type
                        sections.append(current_section)
                        current_section = None
                        continue

                    # Add data to current section
                    if col_b is not None:
                        value = str(col_b).strip() if col_b else None
                        if value and value not in ['Value', '']:
                            current_section[var_name] = self._parse_value(value)

        # Don't forget the last section
        if current_section:
            sections.append(current_section)

        return sections

    def _parse_value(self, value: str) -> Any:
        """
        Parse value string to appropriate type.

        Args:
            value: String value from cell

        Returns:
            Parsed value (bool, int, list, or string)
        """
        # Handle null
        if value.lower() == 'null':
            return None

        # Handle booleans
        if value.lower() in ['true', 'yes']:
            return True
        if value.lower() in ['false', 'no', 'disabled']:
            return False

        # Handle numbers
        try:
            if '.' in value:
                return float(value)
            return int(value)
        except ValueError:
            pass

        # Handle lists (e.g., "[]" or "[item1, item2]")
        if value.startswith('[') and value.endswith(']'):
            inner = value[1:-1].strip()
            if not inner:
                return []
            return [item.strip() for item in inner.split(',')]

        # Return as string
        return value

    def _extract_backend_pools(self) -> List[Dict[str, Any]]:
        """Extract backend address pool configurations."""
        return self._find_section_data('backend_address_pool')

    def _extract_backend_http_settings(self) -> List[Dict[str, Any]]:
        """Extract backend HTTP settings configurations."""
        return self._find_section_data('backend_http_settings')

    def _extract_frontend_ip_configs(self) -> List[Dict[str, Any]]:
        """Extract frontend IP configuration."""
        return self._find_section_data('frontend_ip_configuration')

    def _extract_frontend_ports(self) -> List[Dict[str, Any]]:
        """Extract frontend port configurations."""
        return self._find_section_data('frontend_port')

    def _extract_http_listeners(self) -> List[Dict[str, Any]]:
        """Extract HTTP listener configurations."""
        return self._find_section_data('http_listener')

    def _extract_routing_rules(self) -> List[Dict[str, Any]]:
        """Extract request routing rule configurations."""
        return self._find_section_data('request_routing_rule')

    def _extract_probes(self) -> List[Dict[str, Any]]:
        """Extract probe configurations."""
        return self._find_section_data('probe')

    def get_value(self, variable_name: str) -> Optional[Any]:
        """
        Get a specific value by variable name.

        Args:
            variable_name: The variable name to look up (from Column A)

        Returns:
            The value from Column B, or None if not found
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['APGW']

        for row_idx in range(1, self.ws.max_row + 1):
            col_a = self.ws.cell(row_idx, 1).value

            if col_a and str(col_a).strip() == variable_name:
                col_b = self.ws.cell(row_idx, 2).value
                if col_b and str(col_b).strip() not in ['Value', '']:
                    value = self._parse_value(str(col_b).strip())
                    self.wb.close()
                    return value

        self.wb.close()
        return None
