"""
Build_ENV Sheet Extractor

Extracts data from the Build_ENV sheet which has the structure:
- Column A: Label (human-readable name)
- Column B: Terraform Variable (variable name for terraform)
- Column C: Value (actual value to use)

This sheet contains:
- Project metadata and tags
- Resource Group configuration
- Subscription information
- Application Security Groups
- Subnets
- Private Endpoints
- Key Vault configuration
- User Assigned Identity
- Disk Encryption Set
- Virtual Machine configurations (vm1, vm2, etc.)
"""

import openpyxl
from typing import Dict, Any, List, Optional


class BuildEnvExtractor:
    """Extracts data from Build_ENV sheet."""

    def __init__(self, excel_path: str):
        """
        Initialize the Build_ENV extractor.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.wb = None
        self.ws = None

    def extract(self) -> Dict[str, Any]:
        """
        Extract all data from Build_ENV sheet.

        Returns:
            Dictionary with extracted data organized by section
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['Build_ENV']

        data = {
            'metadata': self._extract_metadata(),
            'tags': self._extract_tags(),
            'subscription': self._extract_subscription(),
            'resource_group': self._extract_resource_group(),
            'location': self._extract_location(),
            'application_security_groups': self._extract_asgs(),
            'subnets': self._extract_subnets(),
            'private_endpoints': self._extract_private_endpoints(),
            'user_assigned_identity': self._extract_user_identity(),
            'key_vault': self._extract_key_vault(),
            'disk_encryption_set': self._extract_disk_encryption(),
            'virtual_machines': self._extract_virtual_machines(),
        }

        self.wb.close()
        return data

    def _get_value_by_tf_var(self, tf_var_name: str, start_row: int = 1, end_row: Optional[int] = None) -> Optional[str]:
        """
        Get value by terraform variable name.

        Searches Column B for the terraform variable name and returns the value from Column C.

        Args:
            tf_var_name: Terraform variable name to search for
            start_row: Row to start searching from
            end_row: Row to end search at (None = end of sheet)

        Returns:
            Value from Column C, or None if not found
        """
        if end_row is None:
            end_row = self.ws.max_row

        for row_idx in range(start_row, end_row + 1):
            col_b = self.ws.cell(row_idx, 2).value  # Column B
            col_c = self.ws.cell(row_idx, 3).value  # Column C

            if col_b and str(col_b).strip() == tf_var_name:
                if col_c and str(col_c).strip() not in ['Value', '']:
                    return str(col_c).strip()

        return None

    def _find_section_start(self, section_name: str) -> Optional[int]:
        """
        Find the row where a section starts.

        Sections are identified by having the section name in Column A
        and 'Terraform Variable' in Column B.

        Args:
            section_name: Name of the section to find

        Returns:
            Row number where section starts, or None if not found
        """
        for row_idx in range(1, self.ws.max_row + 1):
            col_a = self.ws.cell(row_idx, 1).value
            col_b = self.ws.cell(row_idx, 2).value

            if col_a and section_name in str(col_a) and col_b == 'Terraform Variable':
                return row_idx

        return None

    def _extract_metadata(self) -> Dict[str, str]:
        """Extract project metadata (first few rows)."""
        metadata = {}

        # Extract from Overview section
        for row_idx in range(3, 20):
            col_a = self.ws.cell(row_idx, 1).value  # Label
            col_c = self.ws.cell(row_idx, 3).value  # Value

            if col_a and col_c and str(col_c).strip():
                label = str(col_a).strip()
                value = str(col_c).strip()

                # Convert label to snake_case key
                key = label.lower().replace(' ', '_')
                metadata[key] = value

        return metadata

    def _extract_tags(self) -> Dict[str, str]:
        """Extract wab:* tags from the sheet."""
        tags = {}

        # Tags typically start around row 19
        for row_idx in range(19, 35):
            col_b = self.ws.cell(row_idx, 2).value  # Terraform Variable
            col_c = self.ws.cell(row_idx, 3).value  # Value

            if col_b and str(col_b).startswith('wab:'):
                tf_var = str(col_b).strip()
                value = str(col_c).strip() if col_c else None

                if value and value not in ['Value', '']:
                    tags[tf_var] = value

        return tags

    def _extract_subscription(self) -> Optional[str]:
        """Extract subscription value."""
        return self._get_value_by_tf_var('subscription')

    def _extract_resource_group(self) -> Dict[str, str]:
        """Extract resource group configuration."""
        section_start = self._find_section_start('Resource Group')
        if not section_start:
            return {}

        data = {}

        # Search within section (next ~10 rows)
        for row_idx in range(section_start, section_start + 10):
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value != 'Value':
                    data[tf_var] = value

        return data

    def _extract_location(self) -> Optional[str]:
        """Extract location value."""
        return self._get_value_by_tf_var('location')

    def _extract_asgs(self) -> List[Dict[str, str]]:
        """Extract Application Security Groups."""
        asgs = []

        # Find all ASG sections
        row_idx = 1
        while row_idx <= self.ws.max_row:
            col_a = self.ws.cell(row_idx, 1).value
            col_b = self.ws.cell(row_idx, 2).value

            if col_a and 'Application Security Group' in str(col_a) and col_b == 'Terraform Variable':
                # Found an ASG section
                asg_data = {}

                # Read next few rows for ASG data until we hit a blank row or another section
                for offset in range(1, 6):
                    check_row = row_idx + offset
                    label = self.ws.cell(check_row, 1).value
                    tf_var = self.ws.cell(check_row, 2).value
                    value = self.ws.cell(check_row, 3).value

                    # Stop if we hit a blank row (all None)
                    if not label and not tf_var and not value:
                        break

                    # Stop if we hit another section header (Column A has content and Column B = "Terraform Variable")
                    if label and tf_var == 'Terraform Variable':
                        break

                    if tf_var and value and str(value).strip() not in ['Value', '']:
                        asg_data[str(tf_var).strip()] = str(value).strip()

                if asg_data:
                    asgs.append(asg_data)

            row_idx += 1

        return asgs

    def _extract_subnets(self) -> List[Dict[str, str]]:
        """Extract subnet configurations."""
        subnets = []

        # Find subnet section
        section_start = self._find_section_start('Subnet')
        if not section_start:
            return subnets

        subnet_data = {}

        # Read subnet data (next ~10 rows)
        for row_idx in range(section_start, section_start + 15):
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value not in ['Value', 'Terraform Variable']:
                    subnet_data[tf_var] = value

        if subnet_data:
            subnets.append(subnet_data)

        return subnets

    def _extract_private_endpoints(self) -> List[Dict[str, str]]:
        """Extract private endpoint configurations."""
        private_endpoints = []

        # Find private endpoint section
        section_start = self._find_section_start('Private Endpoint')
        if not section_start:
            return private_endpoints

        pe_data = {}

        # Start from row after the header (section_start + 1)
        # Stop at next section header (like "Key Vault")
        for row_idx in range(section_start + 1, section_start + 15):
            col_a = self.ws.cell(row_idx, 1).value
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            # Stop if we hit a new section header (col_a has value AND col_c is "Value")
            if col_a and str(col_c).strip() == 'Value':
                break

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value not in ['Value', 'Terraform Variable']:
                    pe_data[tf_var] = value

        if pe_data:
            private_endpoints.append(pe_data)

        return private_endpoints

    def _extract_user_identity(self) -> Dict[str, str]:
        """Extract user assigned identity configuration."""
        section_start = self._find_section_start('User Assigned Identity')
        if not section_start:
            return {}

        identity_data = {}

        # Read identity data
        for row_idx in range(section_start, section_start + 10):
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value not in ['Value', 'Terraform Variable']:
                    identity_data[tf_var] = value

        return identity_data

    def _extract_key_vault(self) -> Dict[str, str]:
        """Extract key vault configuration."""
        section_start = self._find_section_start('Key Vault')
        if not section_start:
            return {}

        kv_data = {}

        # Read key vault data (including Key Vault Key section)
        for row_idx in range(section_start, section_start + 25):
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value not in ['Value', 'Terraform Variable']:
                    kv_data[tf_var] = value

        return kv_data

    def _extract_disk_encryption(self) -> Dict[str, str]:
        """Extract disk encryption set configuration."""
        section_start = self._find_section_start('Disk Encryption Set')
        if not section_start:
            return {}

        disk_data = {}

        # Read disk encryption data
        for row_idx in range(section_start, section_start + 10):
            col_b = self.ws.cell(row_idx, 2).value
            col_c = self.ws.cell(row_idx, 3).value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                if value and value not in ['Value', 'Terraform Variable']:
                    disk_data[tf_var] = value

        return disk_data

    def _extract_virtual_machines(self) -> List[Dict[str, Any]]:
        """Extract all virtual machine configurations."""
        vms = []

        # Find all VM sections
        row_idx = 1
        while row_idx <= self.ws.max_row:
            col_a = self.ws.cell(row_idx, 1).value
            col_b = self.ws.cell(row_idx, 2).value

            if col_a and 'Virtual Machine' in str(col_a) and col_b == 'Terraform Variable':
                # Found a VM section
                vm_data = {}
                last_data_row = row_idx

                # Read VM data until we hit another section or VM
                for offset in range(1, 50):  # Increased range to capture all data
                    check_row = row_idx + offset
                    if check_row > self.ws.max_row:
                        break

                    label = self.ws.cell(check_row, 1).value
                    tf_var = self.ws.cell(check_row, 2).value
                    value = self.ws.cell(check_row, 3).value

                    # Stop if we hit another section header (but not another VM)
                    if label and tf_var == 'Terraform Variable':
                        # Check if this is another VM section
                        if 'Virtual Machine' in str(label):
                            # Hit next VM section, stop here
                            break
                        else:
                            # Hit a different section, stop here
                            break

                    # Handle special case: Availability Zone (Column A label, Column C value, Column B might be None)
                    if label and str(label).strip() == 'Availability Zone' and value:
                        value_str = str(value).strip()
                        if value_str and value_str not in ['Value', 'None', '']:
                            # Map Zone1, Zone2, Zone3 to 1, 2, 3
                            if value_str.lower().startswith('zone'):
                                zone_num = value_str.lower().replace('zone', '')
                                vm_data['zone'] = zone_num
                            else:
                                vm_data['zone'] = value_str
                            last_data_row = check_row
                        continue

                    if tf_var and value and str(value).strip() not in ['Value', 'None', '']:
                        vm_data[str(tf_var).strip()] = str(value).strip()
                        last_data_row = check_row

                if vm_data:
                    vms.append(vm_data)
                    # Don't skip ahead, just continue from next row to find next VM section
                    row_idx = last_data_row + 1
                    continue

            row_idx += 1

        return vms
