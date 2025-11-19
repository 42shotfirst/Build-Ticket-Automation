"""
NSG Sheet Extractor

Extracts data from the NSG sheet which has a flat table structure:
- Row 1: Headers (name, priority, direction, access, protocol, etc.)
- Row 2+: NSG rule data

All columns should be read as-is to build flat data for terraform.
"""

import openpyxl
from typing import Dict, Any, List, Optional


class NSGExtractor:
    """Extracts network security rules from NSG sheet as flat data."""

    def __init__(self, excel_path: str):
        """
        Initialize the NSG extractor.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.wb = None
        self.ws = None

    def extract(self) -> Dict[str, Any]:
        """
        Extract all NSG rules as flat data.

        Returns:
            Dictionary with NSG rules and metadata
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['NSG']

        data = {
            'rules': self._extract_rules(),
            'headers': self._extract_headers(),
        }

        self.wb.close()
        return data

    def _extract_headers(self) -> List[str]:
        """
        Extract column headers from first row.

        Returns:
            List of header names
        """
        headers = []

        if self.ws.max_row < 1:
            return headers

        # Read first row for headers
        for col_idx in range(1, self.ws.max_column + 1):
            cell_value = self.ws.cell(1, col_idx).value
            if cell_value:
                headers.append(str(cell_value).strip())
            else:
                headers.append(f"column_{col_idx}")

        return headers

    def _extract_rules(self) -> List[Dict[str, Any]]:
        """
        Extract all NSG rules as flat data.

        Reads the entire table structure and converts each row to a dictionary
        using the headers as keys.

        Returns:
            List of rule dictionaries
        """
        rules = []

        if self.ws.max_row < 2:
            return rules

        # Get headers from first row
        headers = self._extract_headers()

        # Read data rows (starting from row 2)
        for row_idx in range(2, self.ws.max_row + 1):
            rule = {}
            has_data = False

            for col_idx, header in enumerate(headers, start=1):
                if col_idx > self.ws.max_column:
                    break

                cell_value = self.ws.cell(row_idx, col_idx).value

                if cell_value is not None and str(cell_value).strip():
                    rule[header] = self._parse_cell_value(str(cell_value).strip())
                    has_data = True
                else:
                    rule[header] = None

            # Only add rule if it has at least some data
            if has_data:
                # Clean up the rule - remove None values for cleaner output
                cleaned_rule = {k: v for k, v in rule.items() if v is not None}
                if cleaned_rule:
                    rules.append(cleaned_rule)

        return rules

    def _parse_cell_value(self, value: str) -> Any:
        """
        Parse cell value and convert to appropriate type.

        Args:
            value: String value from cell

        Returns:
            Parsed value (int, list, bool, or string)
        """
        # Try to convert to int (for priority, port numbers)
        try:
            return int(value)
        except ValueError:
            pass

        # Check for boolean values
        if value.lower() in ['true', 'yes']:
            return True
        if value.lower() in ['false', 'no']:
            return False

        # Check for list values (comma-separated)
        if ',' in value:
            return [item.strip() for item in value.split(',')]

        # Return as string
        return value

    def get_rules_for_terraform(self) -> List[Dict[str, Any]]:
        """
        Get NSG rules formatted for terraform.

        Returns:
            List of rules with terraform-compatible field names
        """
        data = self.extract()
        rules = data.get('rules', [])

        terraform_rules = []

        for idx, rule in enumerate(rules):
            # Map Excel column names to terraform field names
            tf_rule = {
                'name': rule.get('name', f'rule_{idx}'),
                'priority': rule.get('priority', 100 + idx * 10),
                'direction': rule.get('direction', 'Inbound'),
                'access': rule.get('access', 'Allow'),
                'protocol': rule.get('protocol', 'Tcp'),
                'source_port_range': rule.get('source_port_range', '*'),
                'destination_port_ranges': self._parse_port_ranges(rule.get('destination_port_ranges')),
                'source_address_prefix': rule.get('source_address_prefix', '*'),
                'destination_address_prefix': rule.get('destination_address_prefix', '*'),
                'source_asg_keys': self._parse_asg_keys(rule.get('source_asg')),
                'destination_asg_keys': self._parse_asg_keys(rule.get('destination_asg')),
                'description': rule.get('description', f'{rule.get("direction", "Inbound")} {rule.get("access", "Allow")} {rule.get("protocol", "Tcp")} traffic'),
            }

            # Generate source_name and destination_name
            tf_rule['source_name'] = rule.get('source_name', tf_rule['source_asg_keys'][0] if tf_rule['source_asg_keys'] else 'any')
            tf_rule['destination_name'] = rule.get('destination_name', tf_rule['destination_asg_keys'][0] if tf_rule['destination_asg_keys'] else 'any')

            terraform_rules.append(tf_rule)

        return terraform_rules

    def _parse_port_ranges(self, port_value: Any) -> List[str]:
        """
        Parse port ranges into list format.

        Args:
            port_value: Port value (can be string, int, list, or None)

        Returns:
            List of port ranges as strings
        """
        if port_value is None:
            return ['*']

        if isinstance(port_value, list):
            return [str(p) for p in port_value]

        if isinstance(port_value, int):
            return [str(port_value)]

        if isinstance(port_value, str):
            if ',' in port_value:
                return [p.strip() for p in port_value.split(',')]
            return [port_value]

        return ['*']

    def _parse_asg_keys(self, asg_value: Any) -> List[str]:
        """
        Parse ASG keys into list format.

        Args:
            asg_value: ASG value (can be string, list, or None)

        Returns:
            List of ASG keys
        """
        if asg_value is None:
            return []

        if isinstance(asg_value, list):
            return asg_value

        if isinstance(asg_value, str):
            if ',' in asg_value:
                return [a.strip() for a in asg_value.split(',')]
            return [asg_value]

        return []
