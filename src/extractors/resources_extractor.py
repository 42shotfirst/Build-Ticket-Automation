"""
Resources Sheet Extractor

Extracts data from the Resources sheet which has the structure:
- Column A: Label (human-readable name)
- Column B: Terraform Variable (variable name for terraform)
- Column C: Value (actual value to use)

This sheet typically contains additional resource configurations
that supplement or override Build_ENV data.
"""

import openpyxl
from typing import Dict, Any, List, Optional


class ResourcesExtractor:
    """Extracts data from Resources sheet."""

    def __init__(self, excel_path: str):
        """
        Initialize the Resources extractor.

        Args:
            excel_path: Path to the Excel file
        """
        self.excel_path = excel_path
        self.wb = None
        self.ws = None

    def extract(self) -> Dict[str, Any]:
        """
        Extract all data from Resources sheet.

        Returns:
            Dictionary with extracted data organized by terraform variable name
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['Resources']

        data = {
            'terraform_variables': self._extract_all_terraform_variables(),
            'vm_configurations': self._extract_vm_configurations(),
            'tags': self._extract_tags(),
        }

        self.wb.close()
        return data

    def _extract_all_terraform_variables(self) -> Dict[str, str]:
        """
        Extract all terraform variables and their values.

        Reads Column B (terraform variable) and Column C (value) for entire sheet.

        Returns:
            Dictionary mapping terraform variable names to their values
        """
        terraform_vars = {}

        for row_idx in range(1, self.ws.max_row + 1):
            col_b = self.ws.cell(row_idx, 2).value  # Terraform Variable
            col_c = self.ws.cell(row_idx, 3).value  # Value

            if col_b and col_c:
                tf_var = str(col_b).strip()
                value = str(col_c).strip()

                # Skip header rows and empty values
                if value and value not in ['Value', 'Terraform Variable', '']:
                    terraform_vars[tf_var] = value

        return terraform_vars

    def _extract_vm_configurations(self) -> List[Dict[str, Any]]:
        """
        Extract VM configurations from Resources sheet.

        VMs in Resources sheet typically have terraform variables like:
        - vm_list.vm1.name
        - vm_list.vm1.size
        - etc.

        Returns:
            List of VM configuration dictionaries
        """
        vm_data = {}

        for row_idx in range(1, self.ws.max_row + 1):
            col_b = self.ws.cell(row_idx, 2).value  # Terraform Variable
            col_c = self.ws.cell(row_idx, 3).value  # Value

            if col_b and str(col_b).startswith('vm_list.vm'):
                tf_var = str(col_b).strip()
                value = str(col_c).strip() if col_c else None

                if value and value not in ['Value', '']:
                    # Parse vm_list.vm1.property into structured data
                    parts = tf_var.split('.')
                    if len(parts) >= 3:
                        vm_key = parts[1]  # vm1, vm2, etc.
                        property_path = '.'.join(parts[2:])  # name, size, etc.

                        if vm_key not in vm_data:
                            vm_data[vm_key] = {'key': vm_key}

                        # Handle nested properties like tags.role
                        if '.' in property_path:
                            parent, child = property_path.split('.', 1)
                            if parent not in vm_data[vm_key]:
                                vm_data[vm_key][parent] = {}
                            vm_data[vm_key][parent][child] = value
                        else:
                            vm_data[vm_key][property_path] = value

        # Convert to list
        return list(vm_data.values())

    def _extract_tags(self) -> Dict[str, str]:
        """
        Extract wab:* tags from the Resources sheet.

        Returns:
            Dictionary of tag names to values
        """
        tags = {}

        for row_idx in range(1, self.ws.max_row + 1):
            col_b = self.ws.cell(row_idx, 2).value  # Terraform Variable
            col_c = self.ws.cell(row_idx, 3).value  # Value

            if col_b and str(col_b).startswith('wab:'):
                tf_var = str(col_b).strip()
                value = str(col_c).strip() if col_c else None

                if value and value not in ['Value', '']:
                    tags[tf_var] = value

        return tags

    def get_value(self, terraform_variable: str) -> Optional[str]:
        """
        Get a specific value by terraform variable name.

        Args:
            terraform_variable: The terraform variable name to look up

        Returns:
            The value, or None if not found
        """
        self.wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        self.ws = self.wb['Resources']

        for row_idx in range(1, self.ws.max_row + 1):
            col_b = self.ws.cell(row_idx, 2).value

            if col_b and str(col_b).strip() == terraform_variable:
                col_c = self.ws.cell(row_idx, 3).value
                if col_c and str(col_c).strip() not in ['Value', '']:
                    value = str(col_c).strip()
                    self.wb.close()
                    return value

        self.wb.close()
        return None
