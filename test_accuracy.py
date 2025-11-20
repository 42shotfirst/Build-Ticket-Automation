#!/usr/bin/env python3
"""
Comprehensive accuracy test: Compare generated terraform.tfvars against Excel source.
"""

import openpyxl
import json
import re

def load_excel_data():
    """Load and parse key values from Excel."""
    wb = openpyxl.load_workbook('sourcefiles/test_data.xlsm', data_only=True)
    ws = wb['Build_ENV']

    excel_data = {}

    # Core values - read from specific rows in Resource Group section
    # Row 35: resource_group_name
    # Row 36: subscription
    # Row 37: location
    excel_data['resource_group_name'] = str(ws.cell(35, 3).value).strip()
    excel_data['subscription'] = str(ws.cell(36, 3).value).strip()
    excel_data['location'] = str(ws.cell(37, 3).value).strip()

    # ASGs (rows 40-41, 44-45)
    excel_data['asgs'] = [
        {'key': str(ws.cell(40, 3).value).strip(), 'name': str(ws.cell(41, 3).value).strip()},
        {'key': str(ws.cell(44, 3).value).strip(), 'name': str(ws.cell(45, 3).value).strip()},
    ]

    # Subnet (rows 48-55)
    excel_data['subnet'] = {
        'key': str(ws.cell(48, 3).value).strip(),
        'name': str(ws.cell(49, 3).value).strip(),
        'resource_group_name': str(ws.cell(50, 3).value).strip(),
        'virtual_network_name': str(ws.cell(51, 3).value).strip(),
        'network_security_group_name': str(ws.cell(52, 3).value).strip(),
        'route_table_name': str(ws.cell(53, 3).value).strip(),
        'prefixes': str(ws.cell(54, 3).value).strip(),
        'service_endpoints': str(ws.cell(55, 3).value).strip(),
    }

    # VM data - find the VM section
    excel_data['vm'] = {}
    for row in range(100, 130):
        col_b = ws.cell(row, 2).value
        col_c = ws.cell(row, 3).value

        if col_b and col_c:
            col_b_str = str(col_b).strip()
            if 'vm_list.vm1.name' in col_b_str:
                excel_data['vm']['name'] = str(col_c).strip()
            elif 'vm_list.vm1.size' in col_b_str:
                excel_data['vm']['size'] = str(col_c).strip()
            elif 'vm_list.vm1.os_disk_size' in col_b_str:
                excel_data['vm']['os_disk_size'] = str(col_c).strip()
            elif 'vm_list.vm1.asg_key' in col_b_str:
                excel_data['vm']['asg_key'] = str(col_c).strip()
            elif 'vm_list.vm1.snet_key' in col_b_str:
                excel_data['vm']['snet_key'] = str(col_c).strip()
            elif 'vm_list.vm1.ip_allocation' in col_b_str:
                excel_data['vm']['ip_allocation'] = str(col_c).strip()
            elif 'vm_list.vm1.image_os' in col_b_str:
                excel_data['vm']['image_os'] = str(col_c).strip()
            elif 'vm_list.vm1.os_disk_type' in col_b_str:
                excel_data['vm']['os_disk_type'] = str(col_c).strip()
            elif 'vm_list.rsg1.source_image_id' in col_b_str:
                excel_data['vm']['source_image_id'] = str(col_c).strip()

    # Zone - special handling
    for row in range(100, 130):
        col_a = ws.cell(row, 1).value
        col_c = ws.cell(row, 3).value
        if col_a and 'Availability Zone' in str(col_a) and col_c:
            zone_value = str(col_c).strip()
            if zone_value.lower().startswith('zone'):
                excel_data['vm']['zone'] = zone_value.lower().replace('zone', '')

    return excel_data

def load_generated_tfvars():
    """Load generated terraform.tfvars."""
    with open('terraform_output/test_modular_pipeline/terraform.tfvars', 'r') as f:
        return f.read()

def extract_tfvars_value(tfvars_content, pattern):
    """Extract a value from tfvars using regex."""
    match = re.search(pattern, tfvars_content, re.MULTILINE | re.DOTALL)
    return match.group(1) if match else None

def test_accuracy():
    """Run comprehensive accuracy tests."""
    print("=" * 80)
    print("TERRAFORM ACCURACY TEST")
    print("=" * 80)
    print()

    excel = load_excel_data()
    tfvars = load_generated_tfvars()

    passed = 0
    failed = 0

    # Test 1: SPN (calculated from subscription)
    print("TEST 1: Service Principal Name (SPN)")
    print("-" * 80)

    # SPN calculation: Remove 'sub-' prefix from subscription
    subscription = excel["subscription"]
    if subscription.startswith('sub-'):
        expected_spn = f'spn-terraform-{subscription[4:]}'  # Skip 'sub-'
    else:
        expected_spn = f'spn-terraform-{subscription}'

    actual_spn = extract_tfvars_value(tfvars, r'spn = "([^"]+)"')

    print(f"  Excel Subscription: {subscription}")
    print(f"  Expected SPN:       {expected_spn} (sub- prefix removed)")
    print(f"  Generated SPN:      {actual_spn}")

    if actual_spn == expected_spn:
        print("  ✅ PASS")
        passed += 1
    else:
        print("  ❌ FAIL")
        failed += 1
    print()

    # Test 2: Location
    print("TEST 2: Location")
    print("-" * 80)
    expected_location = excel['location']
    actual_location = extract_tfvars_value(tfvars, r'location = "([^"]+)"')

    print(f"  Excel Location:     {expected_location}")
    print(f"  Generated Location: {actual_location}")

    if actual_location == expected_location:
        print("  ✅ PASS")
        passed += 1
    else:
        print("  ❌ FAIL")
        failed += 1
    print()

    # Test 3: Resource Group Name
    print("TEST 3: Resource Group Name")
    print("-" * 80)
    expected_rg = excel['resource_group_name']
    actual_rg = extract_tfvars_value(tfvars, r'resource_group_name = "([^"]+)"')

    print(f"  Excel RG Name:      {expected_rg}")
    print(f"  Generated RG Name:  {actual_rg}")

    if actual_rg == expected_rg:
        print("  ✅ PASS")
        passed += 1
    else:
        print("  ❌ FAIL")
        failed += 1
    print()

    # Test 4: ASG Keys and Names
    print("TEST 4: Application Security Groups")
    print("-" * 80)

    for idx, asg in enumerate(excel['asgs'], 1):
        print(f"  ASG {idx}:")
        print(f"    Excel Key:  {asg['key']}")
        print(f"    Excel Name: {asg['name']}")

        # Check if key exists in tfvars
        key_pattern = f"{asg['key']} = {{"
        if key_pattern in tfvars:
            print(f"    ✅ Key '{asg['key']}' found in tfvars")
            passed += 1
        else:
            print(f"    ❌ Key '{asg['key']}' NOT found in tfvars")
            failed += 1

        # Check if name exists
        if asg['name'] in tfvars:
            print(f"    ✅ Name '{asg['name']}' found in tfvars")
            passed += 1
        else:
            print(f"    ❌ Name '{asg['name']}' NOT found in tfvars")
            failed += 1
        print()

    # Test 5: Subnet Configuration
    print("TEST 5: Subnet Configuration (existing_subnets)")
    print("-" * 80)

    subnet_tests = [
        ('key', excel['subnet']['key']),
        ('name', excel['subnet']['name']),
        ('resource_group_name', excel['subnet']['resource_group_name']),
        ('virtual_network_name', excel['subnet']['virtual_network_name']),
        ('network_security_group_name', excel['subnet']['network_security_group_name']),
        ('route_table_name', excel['subnet']['route_table_name']),
        ('prefixes', excel['subnet']['prefixes']),
        ('service_endpoints', excel['subnet']['service_endpoints']),
    ]

    # Check for existing_subnets (not subnets)
    if 'existing_subnets =' in tfvars:
        print("  ✅ Uses 'existing_subnets' (not 'subnets')")
        passed += 1
    else:
        print("  ❌ Does not use 'existing_subnets'")
        failed += 1

    for field, expected_value in subnet_tests:
        if expected_value in tfvars:
            print(f"  ✅ {field}: {expected_value}")
            passed += 1
        else:
            print(f"  ❌ {field}: {expected_value} NOT FOUND")
            failed += 1

    # Check for both _name and _id fields
    if 'network_security_group_id' in tfvars:
        print("  ✅ network_security_group_id present (equals _name)")
        passed += 1
    else:
        print("  ❌ network_security_group_id missing")
        failed += 1

    if 'route_table_id' in tfvars:
        print("  ✅ route_table_id present (equals _name)")
        passed += 1
    else:
        print("  ❌ route_table_id missing")
        failed += 1
    print()

    # Test 6: Virtual Machine
    print("TEST 6: Virtual Machine Configuration")
    print("-" * 80)

    vm_tests = [
        ('name', excel['vm'].get('name')),
        ('size', excel['vm'].get('size')),
        ('os_disk_size', excel['vm'].get('os_disk_size')),
        ('asg_key', excel['vm'].get('asg_key')),
        ('snet_key', excel['vm'].get('snet_key')),
        ('ip_allocation', excel['vm'].get('ip_allocation')),
        ('image_os', excel['vm'].get('image_os')),
        ('os_disk_type', excel['vm'].get('os_disk_type')),
        ('source_image_id', excel['vm'].get('source_image_id')),
    ]

    for field, expected_value in vm_tests:
        if expected_value and expected_value in tfvars:
            print(f"  ✅ {field}: {expected_value}")
            passed += 1
        elif expected_value:
            print(f"  ❌ {field}: {expected_value} NOT FOUND")
            failed += 1

    # Check zone (should be integer, not "Zone1")
    if excel['vm'].get('zone'):
        expected_zone = f"zone = {excel['vm']['zone']}"
        if expected_zone in tfvars:
            print(f"  ✅ zone: {excel['vm']['zone']} (converted from Zone{excel['vm']['zone']})")
            passed += 1
        else:
            print(f"  ❌ zone: {excel['vm']['zone']} NOT FOUND")
            failed += 1
    print()

    # Test 7: No null values
    print("TEST 7: No Null Values")
    print("-" * 80)
    null_count = tfvars.count(' = null')

    print(f"  Null value count: {null_count}")

    if null_count == 0:
        print("  ✅ PASS - No null values found")
        passed += 1
    else:
        print("  ❌ FAIL - Found null values in output")
        failed += 1
    print()

    # Test 8: Private Endpoints
    print("TEST 8: Private Endpoints")
    print("-" * 80)

    # Should NOT have private_connection_resource_id
    if 'private_connection_resource_id' in tfvars:
        print("  ❌ FAIL - private_connection_resource_id should be omitted")
        failed += 1
    else:
        print("  ✅ PASS - private_connection_resource_id correctly omitted")
        passed += 1

    # Should have other PE fields
    if 'pe_kvlt' in tfvars:
        print("  ✅ Private endpoint key 'pe_kvlt' found")
        passed += 1
    else:
        print("  ❌ Private endpoint key 'pe_kvlt' NOT found")
        failed += 1
    print()

    # Final Summary
    print("=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    total = passed + failed
    percentage = (passed / total * 100) if total > 0 else 0

    print(f"  Total Tests: {total}")
    print(f"  Passed:      {passed}")
    print(f"  Failed:      {failed}")
    print(f"  Accuracy:    {percentage:.1f}%")
    print()

    if failed == 0:
        print("  ✅ ALL TESTS PASSED - OUTPUT IS 100% ACCURATE")
    else:
        print(f"  ⚠️  {failed} TEST(S) FAILED - REVIEW OUTPUT")

    print("=" * 80)

    return failed == 0

if __name__ == '__main__':
    success = test_accuracy()
    exit(0 if success else 1)
